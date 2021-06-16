import pandas as pd
import numpy as np
import xlrd
import os 
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from moviepy.video.io.ffmpeg_tools import ffmpeg_extract_subclip
import matplotlib.pyplot as plt

# first argument arena_xlsx is the original arena in xlsx format
# second argument new_arena_file_name is the file name you want to create in xlsx format
# csv_filename is the name of csv file you want to analyze.
# event_location is either "BR" for Bottom Right, and "TL" for Top Left
# All the arguments must be strings.
# The output is a list:
# 1. the file path of the arena file saved
# 2. a list that contains three lists: 
# time crossed and time exit (in timedelta object), and time crossed (in integer).
def arena_and_event_list(arena_xlsx, new_arena_file_name, csv_filename, event_location):
	workbook = xlrd.open_workbook(arena_xlsx)
	sheet = workbook.sheet_by_name("Sheet1")
	arenaLeft = sheet.cell(1,1).value
	arenaTop = sheet.cell(2,1).value
	arenaWidth = sheet.cell(3,1).value
	arenaHeight = sheet.cell(4,1).value

	floorLeft = sheet.cell(1,2).value
	floorTop = sheet.cell(2,2).value
	floorWidth = sheet.cell(3,2).value
	floorHeight = sheet.cell(4,2).value

	EZLeft = sheet.cell(1,3).value
	EZTop = sheet.cell(2,3).value
	EZWidth = sheet.cell(3,3).value
	EZHeight = sheet.cell(4,3).value

	ECLeft = sheet.cell(1,4).value
	ECTop = sheet.cell(2,4).value
	ECWidth = sheet.cell(3,4).value
	ECHeight = sheet.cell(4,4).value

	FZLeft = sheet.cell(1,5).value
	FZTop = sheet.cell(2,5).value
	FZWidth = sheet.cell(3,5).value
	FZHeight = sheet.cell(4,5).value

	FCLeft = sheet.cell(1,6).value
	FCTop = sheet.cell(2,6).value
	FCWidth = sheet.cell(3,6).value
	FCHeight = sheet.cell(4,6).value

	# Left, Right, Top, Bottom Value for Arena
	arenaL = 0
	arenaR = arenaWidth
	arenaT = arenaHeight
	arenaB = 0

	# Height difference between arena and floor top
	arena_floorT = floorTop - arenaTop

	# Left, Right, Top, Bottom Value for Floor
	floorL = floorLeft - arenaLeft
	floorR = floorL + floorWidth
	floorT = arenaHeight - arena_floorT 
	floorB = floorT - floorHeight

	# Left, Right, Top, Bottom Value for Empty Zone, located at top left corner
	emptyZoneL = floorL
	emptyZoneR = EZWidth
	emptyZoneT = floorT
	emptyZoneB = emptyZoneT - EZHeight

	# Left, Right, Top, Bottom Value for Empty Center, located within the Empty Zone
	emptyCenterL = emptyZoneL
	emptyCenterR = ECWidth
	emptyCenterT = emptyZoneT
	emptyCenterB = emptyCenterT - ECHeight

	# Left, Right, Top, Bottom Value for Food Zone, located at bottom right corner
	foodZoneL = FZLeft - arenaLeft
	foodZoneR = foodZoneL + FZWidth
	foodZoneT = floorB + FZHeight
	foodZoneB = floorB

	# Left, Right, Top, Bottom Value for Food Center, located within the Food Zone
	foodCenterL = foodZoneR - FCWidth
	foodCenterR = foodZoneR
	foodCenterT = foodZoneB + FCHeight
	foodCenterB = floorB

	wb = Workbook()
	ws = wb.active
	ws.title = "Arena"

	# Add values to the new arena file
	ws.append(['Dimension', 'Arena', 'Floor', 'EmptyZone', 'EmptyCenter', 'FoodZone', 'FoodCenter'])
	ws.append(['Left', arenaL, floorL, emptyZoneL, emptyCenterL, foodZoneL, foodCenterL])
	ws.append(['Right', arenaR, floorR, emptyZoneR, emptyCenterR, foodZoneR, foodCenterR])
	ws.append(['Top', arenaT, floorT, emptyZoneT, emptyCenterT, foodZoneT, foodCenterT])
	ws.append(['Bottom', arenaB, floorB, emptyZoneB, emptyCenterB, foodZoneB, foodCenterB])

	# Create an Excel table that starts at A1 and ends on the last non-empty cell
	last_cell = ws.cell(row = ws.max_row, column = ws.max_column).coordinate
	arena_table = Table(displayName = 'ArenaTable', ref = 'A1:{}'.format(last_cell))

	# Style the table and add it to the spreadsheet
	# Note that you can use any of the table styles available in your Excel version (Table Style Medium 6 here) 
	# Finally, save changes.
	style = TableStyleInfo(name = 'TableStyleMedium6', showRowStripes=True)
	arena_table.tableStyleInfo = style
	ws.add_table(arena_table)
	wb.save(new_arena_file_name)
	newarenafilename = os.getcwd()+"/"+new_arena_file_name

	new_df_table = newarena_csv(csv_filename, arenaLeft, arenaTop, arenaHeight)
	if event_location == "BR":
		events = foodzone_events(new_df_table, foodZoneL, foodZoneR, foodZoneB, foodZoneT)
	elif event_location == "TL":
		events = emptyzone_events(new_df_table, emptyZoneL, emptyZoneR, emptyZoneB, emptyZoneT)
	else:
		print("Specify the event_location to either 'BR' or 'TL'")
		return 
	columns_events = events_list(events)
	return newarenafilename, columns_events


def newarena_csv(csv_filename, L, T, H):
    csv_file = pd.read_csv(csv_filename)
    new_columnX, new_columnY, new_columnTime = [],[],[]
    for x in list(csv_file["Body X"]):
        x = x - L
        new_columnX.append(x)
    for y in list(csv_file["Body Y"]):
        y = T + H - y
        new_columnY.append(y)
    format = "%H:%M:%S"
    time_column = csv_file["Time"]
    start_time = time_column.iloc[0][:-3]
    for t in time_column:
        t = t[:-3]
        t = datetime.strptime(t, format) - datetime.strptime(start_time, format)
        new_columnTime.append(t)
    new_df = pd.DataFrame({"Time": new_columnTime,
                         "Body X": new_columnX,
                         "Body Y": new_columnY})
    return new_df

def events_list(new_df_table):
    time_col = new_df_table["Time"]
    j, duration = 0, 0
    time_crossed, time_exit, time_spent = [], [], []
    for i in range(len(time_col) - 1):
        start_time = time_col.iloc[j]
        t_0 = time_col.iloc[i]
        t_1 = time_col.iloc[i + 1]
        t_diff = (t_1 - t_0).total_seconds()
        if t_diff <= 1:
            duration = duration + t_diff
        else:
            time_crossed.append(start_time)
            time_exit.append(t_0)
            time_spent.append(duration)
            j = i + 1
            duration = 0
    return time_crossed, time_exit, time_spent

# Crossing event for empty zone
# It takes in 'data', a table that contains three columns: "Time", "Body X", "Body Y"
def emptyzone_events(data, L, R, B, T):
    t_col, x_col, y_col = [], [], []
    for i, j in data.iterrows():
        t, x, y = j[0], j[1], j[2]
        if x >= L and x <= R and y >= B and y <= T:
            t_col.append(t)
            x_col.append(x)
            y_col.append(y) 
    events = pd.DataFrame({"Time": t_col, "Body X": x_col, "Body Y": y_col})    
    return events

# Crossing event for food zone
# data is a table that contains three columns: "Time", "Body X", "Body Y"
def foodzone_events(data, L, R, B, T):
    t_col, x_col, y_col = [], [], []
    for i, j in data.iterrows():
        t, x, y = j[0], j[1], j[2]
        if x >= L and x <= R and y >= B and y <= T:
            t_col.append(t)
            x_col.append(x)
            y_col.append(y) 
    events = pd.DataFrame({"Time": t_col, "Body X": x_col, "Body Y": y_col})    
    return events

