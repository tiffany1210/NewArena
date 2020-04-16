import pandas as pd 
import numpy as np
import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from moviepy.video.io.ffmpeg_tools import ffmpeg_extract_subclip
import matplotlib.pyplot as plt

# Read data from file 'filename.csv'
csv_file = pd.read_csv("20200220_SUBLAT12-5_Chow.csv")

# Preview the first 5 lines of the loaded data 
csv_file.head()

# Read data from the original Arena file
arena_data = pd.read_excel("Arena.xlsx")

# Open the workbook
workbook = xlrd.open_workbook("Arena.xlsx")

#Open the worksheet
sheet = workbook.sheet_by_name("Sheet1")

#Extract data from the original excel sheet
arenaLeft = sheet.cell(1,1).value
arenaTop = sheet.cell(2,1).value
arenaWidth = sheet.cell(3,1).value
arenaHeight = sheet.cell(4,1).value

floorLeft = sheet.cell(1,2).value
floorTop = sheet.cell(2,2).value
floorWidth = sheet.cell(3,2).value
floorHeight = sheet.cell(4,2).value

EZLeft = sheet.cell(1,3).value
EZTop = sheet.cell(2,3).value
EZWidth = sheet.cell(3,3).value
EZHeight = sheet.cell(4,3).value

ECLeft = sheet.cell(1,4).value
ECTop = sheet.cell(2,4).value
ECWidth = sheet.cell(3,4).value
ECHeight = sheet.cell(4,4).value

FZLeft = sheet.cell(1,5).value
FZTop = sheet.cell(2,5).value
FZWidth = sheet.cell(3,5).value
FZHeight = sheet.cell(4,5).value

FCLeft = sheet.cell(1,6).value
FCTop = sheet.cell(2,6).value
FCWidth = sheet.cell(3,6).value
FCHeight = sheet.cell(4,6).value

# Left, Right, Top, Bottom Value for Arena
arenaL = 0
arenaR = arenaWidth
arenaT = arenaHeight
arenaB = 0

# Height difference between arena and floor top
arena_floorT = floorTop - arenaTop

# Left, Right, Top, Bottom Value for Floor
floorL = floorLeft - arenaLeft
floorR = floorL + floorWidth
floorT = arenaHeight - arena_floorT 
floorB = floorT - floorHeight

# Left, Right, Top, Bottom Value for Empty Zone, located at top left corner
emptyZoneL = floorL
emptyZoneR = EZWidth
emptyZoneT = floorT
emptyZoneB = emptyZoneT - EZHeight

# Left, Right, Top, Bottom Value for Empty Center, located within the Empty Zone
emptyCenterL = emptyZoneL
emptyCenterR = ECWidth
emptyCenterT = emptyZoneT
emptyCenterB = emptyCenterT - ECHeight

# Left, Right, Top, Bottom Value for Food Zone, located at bottom right corner
foodZoneL = FZLeft - arenaLeft
foodZoneR = foodZoneL + FZWidth
foodZoneT = floorB + FZHeight
foodZoneB = floorB

# Left, Right, Top, Bottom Value for Food Center, located within the Food Zone
foodCenterL = foodZoneR - FCWidth
foodCenterR = foodZoneR
foodCenterT = foodZoneB + FCHeight
foodCenterB = floorB

# Open new workbook for new arena file
wb = Workbook()
wb.save("new_arena.xlsx")

# Function for changing a string value to integer or float value
def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value

# Load workbook, select active sheet and rename it
wb = Workbook()
ws = wb.active
ws.title = "Arena"

# Function for changing a string value to integer or float value
def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value

# Add values to the new arena file
ws.append(['Dimension', 'Arena', 'Floor', 'EmptyZone', 'EmptyCenter', 'FoodZone', 'FoodCenter'])
ws.append(['Left', arenaL, floorL, emptyZoneL, emptyCenterL, foodZoneL, foodCenterL])
ws.append(['Right', arenaR, floorR, emptyZoneR, emptyCenterR, foodZoneR, foodCenterR])
ws.append(['Top', arenaT, floorT, emptyZoneT, emptyCenterT, foodZoneT, foodCenterT])
ws.append(['Bottom', arenaB, floorB, emptyZoneB, emptyCenterB, foodZoneB, foodCenterB])

# Create an Excel table that starts at A1 and ends on the last non-empty cell
last_cell = ws.cell(row = ws.max_row, column = ws.max_column).coordinate
arena_table = Table(displayName = 'ArenaTable', ref = 'A1:{}'.format(last_cell))

# Style the table and add it to the spreadsheet
# Note that you can use any of the table styles available in your Excel version (Table Style Medium 6 here) 
# Finally, save changes.
style = TableStyleInfo(name = 'TableStyleMedium6', showRowStripes=True)
arena_table.tableStyleInfo = style
ws.add_table(arena_table)
wb.save("new_arena.xlsx")


# This code gives the new csv file that has been modified according to the new arena setting.
# The code below reformats the x body of the mouse.
def reformatX(column):
    new_column = []
    for x in list(column):
        x = x - arenaLeft
        new_column.append(x)
    return new_column

# This reformats the y body of the mouse.
def reformatY(column):
    new_column = []
    for y in list(column):
        y = arenaTop + arenaHeight - y
        new_column.append(y)
    return new_column

# This reformats the timestamp - it resets the starting time to 00:00:00.
# The new column is in datetime object
def time_to_num(time_column):
    new_column = []
    format = "%H:%M:%S"
    start_time = time_column.iloc[0][:-3]
    for t in time_column:
        t = t[:-3]
        t = datetime.strptime(t, format) - datetime.strptime(start_time, format)
        new_column.append(t)
    return new_column

#This will now reformat the given csv_file.
# new_df is the data that we will use, which is a dataframe table.
new_columnX = reformatX(csv_file["Body X"])
new_columnY = reformatY(csv_file["Body Y"])
new_columnTime = time_to_num(csv_file["Time"])
new_df = pd.DataFrame({"Time": new_columnTime, "Body X": new_columnX, "Body Y": new_columnY})

new_df.to_csv("new_20200220_SUBLAT12-5_Chow.csv", index=False)


# Crossing event for empty zone
# It takes in 'data', a table that contains three columns: "Time", "Body X", "Body Y"
def emptyzone_events(data):
    t_col, x_col, y_col = [], [], []
    for i, j in data.iterrows():
        t, x, y = j[0], j[1], j[2]
        if x >= emptyZoneL and x <= emptyZoneR and y >= emptyZoneB and y <= emptyZoneT:
            t_col.append(t)
            x_col.append(x)
            y_col.append(y) 
    events = pd.DataFrame({"Time": t_col, "Body X": x_col, "Body Y": y_col})    
    return events

# Crossing event for food zone
# data is a table that contains three columns: "Time", "Body X", "Body Y"
def foodzone_events(data):
    t_col, x_col, y_col = [], [], []
    for i, j in data.iterrows():
        t, x, y = j[0], j[1], j[2]
        if x >= foodZoneL and x <= foodZoneR and y >= foodZoneB and y <= foodZoneT:
            t_col.append(t)
            x_col.append(x)
            y_col.append(y) 
    events = pd.DataFrame({"Time": t_col, "Body X": x_col, "Body Y": y_col})    
    return events

# A dataframe table for crossing events with columns "Time", "Body X", "Body Y"
empty_events = emptyzone_events(new_df)
fz_events = foodzone_events(new_df)

# This provides an events table with the event number (natural number), 
# time crossed and time exit (datetime object), and time spent (integer).
# It takes in time column of the table from emptyzone_events and foodzone_events.
def event_time(time_col):
    j, event_num, duration = 0, 1, 0
    event_num_col, time_crossed, time_exit, time_spent = [], [], [], []
    for i in range(len(time_col) - 1):
        start_time = time_col.iloc[j]
        t_0 = time_col.iloc[i]
        t_1 = time_col.iloc[i + 1]
        t_diff = (t_1 - t_0).total_seconds()
        if t_diff <= 1:
            duration = duration + t_diff
        else:
            event_num_col.append(event_num)
            time_crossed.append(start_time)
            time_exit.append(t_0)
            time_spent.append(duration)
            j = i + 1
            event_num += 1
            duration = 0
    events = pd.DataFrame({"Event number": event_num_col, "Time crossed": time_crossed,
                          "Time exit": time_exit, "Time spent": time_spent})
    return events

# A dataframe table for crossing events with columns "Event number", "Time crossed", "Time exit", and "Time spent"
pd_df = event_time(empty_events["Time"])
pd_df_2 = event_time(fz_events["Time"])


# 1. event_table is a dataframe table of crossing events with columns
# Event number", "Time crossed", "Time exit", and "Time spent"(pd_df, pd_df_2)
# 2. from_video is a string of the name of the video to extract from.
# e.g. "20200220_SUBLAT12-5_Jelly-Exposure (Converted).mov"
# 3. to_video_name is a string of the name you want the extracted video to be.
# e.g. "test" would give out "test1.mov", "test2.mov", and so on.
# The function returns nothing, but produces extracted video clips of crossing events from the event table.
# The number of videos would correspond to the number of events specified in the event_table.
def extract_video(event_table, from_video, to_video_name):
    time_crossed_col = event_table["Time crossed"]
    time_exit_col = event_table["Time exit"]
    for i in range(len(event_table)):
        t1 = time_crossed_col.iloc[i].total_seconds()
        t2 = time_exit_col.iloc[i].total_seconds()
        ffmpeg_extract_subclip(from_video, t1, t2, to_video_name + str(i+1)+".mov")
    return

# It plots a histogram of the crossing events
# x axis is labelled as the event number and the y axis is the Duration (Seconds).
# pd_df can be replaced with any dataframe event table.
# The title of the histogram and the savefig title should correspond to the specific event table.
number_events = np.arange(len(pd_df.index))
plt.bar(number_events, pd_df['Time spent (seconds)'])
plt.xticks(number_events, pd_df['Event number'])
plt.ylabel("Duration (Seconds)")
plt.xlabel("Event number")
plt.title("Event for Empty Zone (20200220_SUBLAT12-5_Jelly-Exposure)")
plt.savefig('Event for Empty Zone.jpg')


# It generates an excel sheet with a given sheet_name
# sheet_name should be in string
# dataframe is a dataframe table of crossing events
# excel_loc is the location of the excel. e.g."20200220_SUBLAT12-5_Jelly-Exposure.xlsx"
def generate_excel(dataframe, excel_loc, sheet_name):
    writer = pd.ExcelWriter(excel_loc, engine = 'openpyxl')
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    return(writer.close())


# Table of distance from the empty center 
# Takes in data, which is a dataframe table with columns "Time", "Body X", and "Body Y"
def distance_from_empty_center(data):
    t_col, dist_col = [], []
    x_center = (emptyCenterR + emptyCenterL)/2
    y_center = (emptyCenterT + emptyCenterB)/2
    for i, j in data.iterrows():
        t, x, y = j[0], j[1], j[2]
        distance = ((x - x_center)**2 +(y - y_center))**(1/2)
        t_col.append(t)
        dist_col.append(distance)
    events = pd.DataFrame({"Time": t_col, "Distance from empty center": dist_col})    
    return events


# Table of distance from the food center
def distance_from_food_center(data):
    t_col, dist_col = [], []
    x_center = (foodCenterR + foodCenterL)/2
    y_center = (foodCenterT + foodCenterB)/2
    for i, j in data.iterrows():
        t, x, y = j[0], j[1], j[2]
        distance = ((x - x_center)**2 +(y - y_center))**(1/2)
        t_col.append(t)
        dist_col.append(distance)
    events = pd.DataFrame({"Time": t_col, "Distance from food center": dist_col})    
    return events

# Creates a distance table from a dataframe table
distance_tbl = distance_from_empty_center(new_df)
distance_tbl_2 = distance_from_food_center(new_df)


# It writes an excel sheet that contains
# 1. Empty Zone event table 
# 2. Food Zone event table
# 3. Distance from Empty center
# 4. Distance from Food center
# The dataframe table and sheet name should change accordingly.
with pd.ExcelWriter("20200220_SUBLAT12-5_Jelly-Exposure.xlsx", 
                        engine = 'openpyxl') as writer:
    
    pd_df.to_excel(writer, sheet_name="Empty Zone event table", index=False)
    pd_df_2.to_excel(writer, sheet_name="Food Zone event table", index=False)
    distance_tbl.to_excel(writer, sheet_name = "Distance from Empty Center")
    distance_tbl_2.to_excel(writer, sheet_name = "Distance from Food Center")



